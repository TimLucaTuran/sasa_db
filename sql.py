import mysql.connector as mc
import os, sys
import openpyxl

mydb = mc.connect(
    host = "localhost",
    user = "root",
    passwd = "tltit69z",
    database = "meta_materials",
)
my_cursor = mydb.cursor()
wb = openpyxl.load_workbook("project_overview.xlsx")
sheets = wb.sheetnames
ws = wb[sheets[int(sys.argv[1]) - 1]] #usable sheets are: 10,

my_cursor.execute("TRUNCATE TABLE simulations;")
my_cursor.execute("TRUNCATE TABLE wire;")
my_cursor.execute("TRUNCATE TABLE square;")
my_cursor.execute("TRUNCATE TABLE circ;")
my_cursor.execute("TRUNCATE TABLE disc;")


sql_dict={ 'geometry': None ,
           'm_file': None ,
           'particle_material': None ,
           'cladding': None ,
           'substrate': None ,
           'periode': None ,
           'wavelength_start': None ,
           'wavelength_stop': None ,
           'spectral_points': None ,
           'simulation_order': None ,
           'adress': None ,
           'length': None ,
           'width': None ,
           'thickness': None ,
           'corner_radius': None ,
           'radius': None ,
           'rounded_corner': None ,
           'hole': None,
           'simulation_id': None ,
          }


class Exl:
#Define the Excel-Struct it holds information of a Excel-Cell and which
#opperations need to be applied. For example the wavelength excel cell:
#"0.64 ... 1.7" needs to be split into wavelength_start and _stop
    def __init__(self, name, target_key, opperation_list=[]):
        self.name = name
        self.target_key = target_key
        self.opperation_list = opperation_list
        self.column = None
        self.data = None
        self.target_dict = sql_dict
        self.geometries_list = ['wire', 'square', 'circ']

    def write(self):
        self.target_dict[self.target_key] = self.data
        return


    def wav_split(self):
        if '…' in self.data:
            both = self.data.split('…')
        elif '-' in self.data:
            both = self.data.split('-')
        elif '–' in self.data:
            both = self.data.split('–')

        else:
            print("couldn't wav-split")
            return
        start = both[0]
        stop = both[1]

        #copy stop to the SQL list, start will be taken care of by write()
        self.target_dict['wavelength_stop'] = stop
        self.data = start
        return

    def evaluate(self):
        try:
            self.data = float(eval(self.data))
        except:
            pass
        return

    def listify(self):
        """
        Some Excel Cells have Values like [5 7 9 11] or matlab like
        25:5:75. This method needs to turn these into python lists.
        """
        if type(self.data) is str:
            if ':' in self.data:
                nums = self.data.split(':')
                nums = [int(i) for i in nums]
                if len(nums) == 3:
                    start = nums[0]
                    step = nums[1]
                    stop = nums[2]
                elif len(nums) == 2:
                    start = nums[0]
                    step = 1
                    stop = nums[1]
                else:
                    raise ValueError('Weird number of ":" ')

                self.data = list(range(start, stop+step, step))

            elif '[' in self.data:
                self.data = self.data.replace('[', '').replace(']', '')
                self.data = self.data.replace(',', ' ').split()

            elif self.data.count(',') > 1:
                self.data = self.data.split(',')

            elif ', ' in self.data:
                self.data = self.data.split(', ')

        return

    def comma_split(self):
        if type(self.data) is str:
            if ',' in self.data:
                self.data = self.data.split(',')
        return

    def geo_setup(self):
        self.data = self.data.lower()

        #check for additional geometry features
        if 'rounded' in self.data:
            self.target_dict['rounded_corner'] = 1

        if 'hole' in self.data:
            self.target_dict['hole'] = 1

        if 'sem' in self.data:
            raise RuntimeError('Have not implemented SEM wires jet')
            return
        #check for valid sql-geometry
        for geo in self.geometries_list:
            if geo in self.data:
                self.data = geo
                break
        else:
            raise RuntimeError('Found no valid geometry in : {}'.format(self.data))
        return




class QueryGenerator:
    def __init__(self, target_dict = sql_dict):
        self.target_dict = target_dict
        self.sim_query = 'INSERT INTO simulations ('
        self.geo_query = 'INSERT INTO '
        self.valid_queries = 0
        self.failed_queries = 0
        self.skip_row = False
        self.wire = ['simulation_id', 'length', 'width', 'thickness','corner_radius',
                    'rounded_corner']
        self.square = ['simulation_id', 'length', 'width', 'thickness', 'hole']
        self.circ = ['simulation_id','width', 'thickness', 'hole']
        self.geometries = {'wire' : self.wire,
                           'square' : self.square,
                           'circ' : self.circ,
                           }
        self.adress = []

    def update_id(self):
        #get the current simulation_id
        my_cursor.execute("SELECT MAX(simulation_id) FROM simulations;")
        id = my_cursor.fetchone()[0]
        if id is None:
            id = 1
        else:
            id += 1
        print("ID: ", id)
        self.target_dict['simulation_id'] = id
        return

    def reset(self):
        self.sim_query = 'INSERT INTO simulations ('
        self.geo_query = 'INSERT INTO '
        self.target_dict['hole'] = None
        self.target_dict['rounded_corner'] = None
        return

    def make_query(self, sql_dict):
        ###simulations query
        #consruct the list which holds the sql data to be submitted
        sim_data = []
        key_count = 0
        for key, val in sql_dict.items():
            if key == 'length':  #length is the first key for the geo-query
                break
            elif val is not None:
                sim_data.append(val)
                self.sim_query += key
                self.sim_query += ', '
                key_count += 1

        self.sim_query = self.sim_query[:-2]
        self.sim_query += ') VALUES ('
        self.sim_query += len(sim_data)* "%s, "
        self.sim_query = self.sim_query[:-2]
        self.sim_query += ')'

        #geometry query
        self.update_id()
        current_geo = self.geometries[sql_dict['geometry']]
        self.geo_query += sql_dict['geometry']
        self.geo_query += ' ('

        #select the activ part of the dict for current geometry
        active_dict = [(key, val) for key, val in sql_dict.items() if key in current_geo]

        geo_data = []
        for key, val in active_dict:
            if sql_dict[key] is not None:
                geo_data.append(val)
                self.geo_query += key
                self.geo_query += ', '

        self.geo_query = self.geo_query[:-2]
        self.geo_query += ') VALUES ('
        self.geo_query += len(geo_data)*"%s, "
        self.geo_query = self.geo_query[:-2]
        self.geo_query += ')'

        #Try to execute the queries
        try:
            my_cursor.execute(self.sim_query, tuple(sim_data))
            my_cursor.execute(self.geo_query, tuple(geo_data))
            self.valid_queries += 1
        except Exception as e:
            self.failed_queries += 1
            print('Bad query')
            print(e)
            print(self.sim_query)
            print(sim_data)
            print(self.geo_query)
            print(geo_data)
            print('\n')


        self.reset()

        return

    def generate(self, sql_dict):
        """
        This method gets a possibly nested sql_dict.
        It splits them up and calls make_query for every simple list.
        I thought about a recursiv way to write this function but I can't get
        it to work.
        """
        #find the colums which have list-values
        list_vals = [(key, val) for key, val in sql_dict.items()
                    if type(val) is list]
        list_vals.sort(key = lambda tup : len(tup[1]), reverse = True)

        #check if there are multiple values with the same dimension.
        #In this case the adress system won't work because it can't know which
        #dimension belongs to which property
        len_list = [len(tup[1]) for tup in list_vals]
        if len(len_list) != len(set(len_list)):
            self.failed_queries += 1
            print('multiple Values have the same dimension, skipping')
            return

        #reset the adress
        self.adress = []


        if len(list_vals) == 0:
            self.make_query(sql_dict)
            return

        elif len(list_vals) == 1:
            current_key = list_vals[0][0]
            current_list = list_vals[0][1]
            self.adress.append(None)

            for i in range(len(current_list)):
                sql_dict[current_key] = current_list[i]
                self.adress[-1] = i
                sql_dict['adress'] = str(self.adress)
                self.make_query(sql_dict)


            #reset the adress
            sql_dict['adress'] = None
            return

        elif len(list_vals) == 2:
            key_0 = list_vals[0][0]
            list_0 = list_vals[0][1]
            key_1 = list_vals[1][0]
            list_1 = list_vals[1][1]
            self.adress.append(None)
            self.adress.append(None)

            for i in range(len(list_0)):
                sql_dict[key_0] = list_0[i]
                self.adress[0] = i
                for j in range(len(list_1)):
                    sql_dict[key_1] = list_1[j]
                    self.adress[1] = j
                    sql_dict['adress'] = str(self.adress)
                    self.make_query(sql_dict)

            #reset the adress
            sql_dict['adress'] = None
            return

        elif len(list_vals) == 3:
            key_0 = list_vals[0][0]
            list_0 = list_vals[0][1]
            key_1 = list_vals[1][0]
            list_1 = list_vals[1][1]
            key_2  = list_vals[2][0]
            list_2 = list_vals[2][1]
            self.adress.append(None)
            self.adress.append(None)
            self.adress.append(None)

            for i in range(len(list_0)):
                sql_dict[key_0] = list_0[i]
                self.adress[0] = i
                for j in range(len(list_1)):
                    sql_dict[key_1] = list_1[j]
                    self.adress[1] = j
                    for k in range(len(list_2)):
                        sql_dict[key_2] = list_1[k]
                        self.adress[2] = k
                        sql_dict['adress'] = str(self.adress)
                        self.make_query(sql_dict)

            #reset the adress
            sql_dict['adress'] = None
            return

        else:
            self.failed_queries += 1
            print('list_val dim to high')
            print('list_vals: ',list_vals)



#Define the excel_list with one Exl for every column
#Setup ist Exl(excel_name, sql_name, [opperations])
exl_list = [Exl('m-file', 'm_file'),
            Exl('material', 'particle_material', [Exl.comma_split]),
            Exl('cladding', 'cladding', [Exl.comma_split]),
            Exl('substrate', 'substrate',  [Exl.comma_split]),
            Exl('geom', 'geometry', [Exl.geo_setup]),
            Exl('periode', 'periode', [Exl.listify] ),
            Exl('wavelength', 'wavelength_start', [Exl.wav_split]),
            Exl('points', 'spectral_points'),
            Exl('order', 'simulation_order', [Exl.listify]),
            Exl('length', 'length', [Exl.evaluate, Exl.listify]),
            Exl('width', 'width', [Exl.evaluate, Exl.listify]),
            Exl('thickness', 'thickness', [Exl.evaluate, Exl.listify]),
            Exl('corner radius', 'corner_radius', [Exl.listify]),
            ]


####Read the excel data into the exl_list
#find the row with the column names
for row in ws.iter_rows():
    if 'file' in row[0].value:
        name_row = row

#find the position of each excel collumn
for cell in name_row:
    for exl in exl_list:
        if exl.name in cell.value and exl.column is None:
            exl.column = cell.column -1

####Main loop: Generate SQL-queries for every Excel row####
query_gen = QueryGenerator()
for row in ws.iter_rows(name_row[0].row + 1, ws.max_row): #ws.max_row):
    #Break on empty row
    if len(row[0].value) == 0:
        break
    #Load the data of the row
    print('row: ', row[0].row)
    for i in range(len(row)):
        for exl in exl_list:
            if i == exl.column:
                exl.data = row[i].value

    #Construct the sql_dict out of the exl_list
    for exl in exl_list:
        #print(exl.name, 'at', exl.column)
        for opperation in exl.opperation_list:
            try:
                opperation(exl)
            except Exception as e:
                print(e)
                query_gen.skip_row = True
        exl.write()

    if not query_gen.skip_row:
        query_gen.generate(sql_dict)
    else:
        print('skipping row: ', row[0].row)
        query_gen.failed_queries += 1
        query_gen.skip_row = False
    print('')


print("{}/{} queries successful".format(
query_gen.valid_queries, query_gen.valid_queries + query_gen.failed_queries))
commit = input('Commit changes?')
if commit == 'y':
    mydb.commit()
