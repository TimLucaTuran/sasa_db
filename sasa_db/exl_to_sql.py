import sqlite3
import os, sys
import openpyxl
import argparse


class Exl:
    """Defines the Excel-Struct it holds information of a single Excel-Cell and which
    opperations need to be applied. E.g. for the wavelength excel cell
    "0.64 ... 1.7" needs to be split into wavelength_start and wavelength_stop
    """


    target_dict = None #This defines a shared class attribute
    skip_list = None

    def __init__(self, name, target_key, opperation_list=[]):
        self.name = name
        self.target_key = target_key
        self.opperation_list = opperation_list
        self.column = None
        self.data = None
        self.target_dict = sql_dict
        self.geometries_list = ['wire', 'square', 'circ']

    def write(self):
        self.target_dict[self.target_key] = self.data
        return


    def wav_split(self):
        if '…' in self.data:
            both = self.data.split('…')
        elif '-' in self.data:
            both = self.data.split('-')
        elif '–' in self.data:
            both = self.data.split('–')

        else:
            print("couldn't wav-split")
            return
        start = both[0].strip()
        stop = both[1].strip()

        #copy stop to the SQL list, start will be taken care of by write()
        self.target_dict['wavelength_stop'] = stop
        self.data = start
        return

    def evaluate(self):
        try:
            self.data = float(eval(self.data))
        except:
            pass
        return

    def listify(self):
        """
        Some Excel Cells have Values like [5 7 9 11] or matlab like
        25:5:75. This method needs to turn these into python lists.
        """
        if type(self.data) is str:
            if ':' in self.data:
                nums = self.data.split(':')
                nums = [int(i) for i in nums]
                if len(nums) == 3:
                    start = nums[0]
                    step = nums[1]
                    stop = nums[2]
                elif len(nums) == 2:
                    start = nums[0]
                    step = 1
                    stop = nums[1]
                else:
                    raise ValueError('Weird number of ":" ')

                self.data = list(range(start, stop+step, step))

            elif '[' in self.data:
                self.data = self.data.replace('[', '').replace(']', '')
                self.data = self.data.replace(',', ' ').split()

            elif self.data.count(',') > 1:
                self.data = self.data.split(',')

            elif ', ' in self.data:
                self.data = self.data.split(', ')

        return

    def comma_split(self):
        if type(self.data) is str:
            if ',' in self.data:
                self.data = self.data.split(',')
        return

    def trim(self):
        for d in self.data:
            d = d.strip()

    def sem_check(self):
        if not self.target_dict['image_source'] == 'SEM_FLAG':
            self.data = None
        return

    def geo_setup(self):
        self.data = self.data.lower()

        #check for additional geometry features
        if 'rounded' in self.data:
            self.target_dict['rounded_corner'] = 1

        if 'hole' in self.data:
            self.target_dict['hole'] = 'holes'

        if 'sem' in self.data:
            self.target_dict['image_source'] = 'SEM_FLAG'


        #check for valid sql-geometry
        for geo in self.geometries_list:
            if geo in self.data:
                self.data = geo
                break
        else:
            raise RuntimeError('Found no valid geometry in : {}'.format(self.data))
        return

    def skip_check(self):
        #skip manually marked rows
        if 'skip' in self.data:
            raise ValueError('skipping marked row')

        #skip existing
        if args["skip_existing"]:
            if self.data in self.skip_list:
                raise ValueError('skipping existing row', self.data)
        return


class QueryGenerator:
    def __init__(self, target_dict):
        self.target_dict = target_dict
        self.sim_query = 'INSERT INTO simulations ('
        self.geo_query = 'INSERT INTO '
        self.valid_queries = 0
        self.failed_queries = 0
        self.skip_row = False
        self.wire = ['simulation_id', 'length', 'width', 'thickness', 'hole',
                    'corner_radius', 'rounded_corner', 'image_source']
        self.square = ['simulation_id', 'width', 'thickness', 'hole']
        self.circ = ['simulation_id', 'width', 'thickness', 'hole']
        self.L = []
        self.geometries = {'wire' : self.wire,
                           'square' : self.square,
                           'circ' : self.circ,
                           'L' : self.L,
                           }
        self.dummy_dict = {'Chi_RotWire_1_rounded_Ti_d' : 0,
                           'Chi_RotWire_1_rounded_Ti_h' : 0,
                           'Chi_RotWire_1_rounded_Ti_k' : 0,
                           'Chi_RotWire_2_rounded_Ti_a' : 0,
                           'Chi_RotWire_2_rounded_Ti_g' : 0,
                           }

    def update_id(self):
        #get the current simulation_id
        my_cursor.execute("SELECT last_insert_rowid()")
        id = my_cursor.fetchone()[0]
        print("ID: ", id)
        self.target_dict['simulation_id'] = id
        return

    def reset(self):
        self.target_dict['hole'] = None
        self.target_dict['rounded_corner'] = None
        return

    def dummy_dimension_check(self, sql_dict):
        """Some Data-Entries have extra dimensions which are not recorded
        for example: Ti Adhesion-Layer-Height. These need to be squished."""
        name = sql_dict['m_file']
        #print(name, sql_dict['adress'], type(sql_dict['adress']))
        if name in self.dummy_dict:
            #modify the adress so that only dummy attribute 0 can be accesed
            if sql_dict['adress'] is None:
                sql_dict['adress'] = [0]
            else:
                eval(sql_dict['adress']).insert(self.dummy_dict[name], 0)
        return



    def make_query(self, sql_dict):
        self.sim_query = 'INSERT INTO simulations ('
        self.geo_query = 'INSERT INTO '

        self.dummy_dimension_check(sql_dict)

        ###simulations query
        #construct the list which holds the sql data to be submitted
        sim_data = []
        key_count = 0
        for key, val in sql_dict.items():
            #width is the first key for the geo-query. This is pretty dirty
            #because different parameters might cause a fault here
            if key == 'width':
                break
            elif val is not None:
                sim_data.append(str(val))
                self.sim_query += key
                self.sim_query += ', '
                key_count += 1

        self.sim_query = self.sim_query[:-2]
        self.sim_query += ') VALUES ('
        self.sim_query += len(sim_data)* "?, "
        self.sim_query = self.sim_query[:-2]
        self.sim_query += ')'

        try:
            with conn:
                my_cursor.execute(self.sim_query, tuple(sim_data))

        except Exception as e:
            self.failed_queries += 1
            print('Bad query')
            print(e)
            print(self.sim_query)
            print(sim_data)
            return

        #geometry query
        self.update_id()
        current_geo = self.geometries[sql_dict['geometry']]
        self.geo_query += sql_dict['geometry']
        self.geo_query += ' ('


        #select the activ part of the dict for current geometry
        active_dict = [(key, val) for key, val in sql_dict.items() if key in current_geo]

        geo_data = []
        for key, val in active_dict:
            if sql_dict[key] is not None:
                geo_data.append(str(val))
                self.geo_query += key
                self.geo_query += ', '

        self.geo_query = self.geo_query[:-2]
        self.geo_query += ') VALUES ('
        self.geo_query += len(geo_data)*"?, "
        self.geo_query = self.geo_query[:-2]
        self.geo_query += ')'

        #Try to execute the queries
        try:
            with conn:
                my_cursor.execute(self.geo_query, tuple(geo_data))
            self.valid_queries += 1

        except Exception as e:
            print('Bad query:')
            print(e)
            print(self.geo_query)
            print(geo_data)
            print('\n')
            self.failed_queries += 1




        return

    def generate(self, sql_dict):
        """
        This method gets a possibly nested sql_dict.
        It splits them up and calls make_query for every simple list.
        I thought about a recursiv way to write this function but I can't get
        it to work.
        """
        #find the colums which have list-values
        list_cols = [(key, val) for key, val in sql_dict.items()
                    if type(val) is list]
        list_cols.sort(key = lambda tup : len(tup[1]))

        #if the sql_dict contains no lists just make the query
        if len(list_cols) == 0:
            self.make_query(sql_dict)
            return

        #check for list of the same length and bring them in the correct order
        len_list = [len(tup[1]) for tup in list_cols]




        """
        #check if there are multiple values with the same dimension.
        #In this case the adress system won't work because it can't know which
        #dimension belongs to which property
        len_list = [len(tup[1]) for tup in list_cols]
        if len(len_list) != len(set(len_list)):
            self.failed_queries += 1
            print('multiple Values have the same dimension, skipping')
            return
        """

        #prepare the adress
        sql_dict['adress'] = [0]*len(list_cols)
        #make queries
        self.generate_split(list_cols, sql_dict)
        #reset the adress
        sql_dict['adress'] = None

    def generate_split(self, list_cols, sql_dict, idx=0):
        if len(list_cols) == idx:
            #Here sql_dict contains no more lists
            self.make_query(sql_dict)

        else:
            #Here there are more list-type-values wich need to be split up
            current_key, current_list = list_cols[idx]

            for i in range(len(current_list)):
                sql_dict['adress'][idx] = i
                sql_dict[current_key] = current_list[i]
                self.generate_split(list_cols, sql_dict, idx=idx+1)

# construct the argument parse and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("excel-sheet", metavar="exl", help="path to excel-file", default="NN_smats.xlsx")
ap.add_argument("databank", metavar="db", help="path to sqlite3-db", default="NN_smats.db")
ap.add_argument("-n", "--sheet-number", help="which excel-sheet to convert",
    type=int, default=1)
ap.add_argument("-v", "--verbose", action="store_true", help="verbose output")
ap.add_argument("-s", "--skip-existing", action="store_true", help="skipping rows already contained in the db")
args = vars(ap.parse_args())


#Connect to the sqlite3 db and the exel sheet
conn = sqlite3.connect(args["databank"])
my_cursor = conn.cursor()
wb = openpyxl.load_workbook(args["excel_sheet"])
sheets = wb.sheetnames
sheet_number = args["sheet_number"]
ws = wb[sheets[sheet_number - 1]] #usable sheets are: 10, 11, 14, 16, 22

sql_dict={ 'geometry': None ,
           'm_file': None ,
           'particle_material': None ,
           'cladding': None ,
           'substrate': None ,
           'periode': None ,
           'wavelength_start': None ,
           'wavelength_stop': None ,
           'spectral_points': None ,
           'simulation_order': None ,
           'adress': None ,
           'angle_of_incidence' : None,
           'width': None ,
           'length': None ,
           'thickness': None ,
           'corner_radius': None ,
           'radius': None ,
           'rounded_corner': None ,
           'hole': None,
           'girth': None,
           'image_source': None,
           'simulation_id': None ,
          }

Exl.target_dict = sql_dict

#Define the excel_list with one Exl for every column
#Setup ist Exl(excel_name, sql_name, [opperations])
exl_list = [Exl('m-file', 'm_file', [Exl.skip_check]),
            Exl('particle material', 'particle_material', [Exl.comma_split, Exl.trim]),
            Exl('cladding', 'cladding', [Exl.comma_split]),
            Exl('substrate', 'substrate',  [Exl.comma_split]),
            Exl('geometry', 'geometry', [Exl.geo_setup]),
            Exl('periode', 'periode', [Exl.listify] ),
            Exl('wavelength', 'wavelength_start', [Exl.wav_split]),
            Exl('spectral points', 'spectral_points'),
            Exl('order', 'simulation_order', [Exl.listify]),
            Exl('angle of incidence', 'angle_of_incidence', [Exl.listify]),
            Exl('length', 'length', [Exl.evaluate, Exl.listify]),
            Exl('width', 'width', [Exl.evaluate, Exl.listify]),
            Exl('thickness', 'thickness', [Exl.evaluate, Exl.listify]),
            Exl('corner radius', 'corner_radius', [Exl.listify]),
            Exl('draw file', 'image_source', [Exl.sem_check]),
            Exl('girth', 'girth', [Exl.listify])
            ]


####Read the excel data into the exl_list
#find the row with the column names
for row in ws.iter_rows():
    if 'file' in row[0].value:
        name_row = row
        break


#find the position of each excel collumn
for cell in name_row:
    if cell.value is None:
        break

    for exl in exl_list:
        if exl.name == cell.value.strip():
            exl.column = cell.column - 1
            break

if args["verbose"]:
    for exl in exl_list:
        print(exl.name, exl.column)

if args["skip_existing"]:
    #query the db for already existing entries
    query = """SELECT DISTINCT m_file FROM simulations"""
    my_cursor.execute(query)
    Exl.skip_list = [x[0] for x in my_cursor.fetchall()]

####Main loop: Generate SQL-queries for every Excel row####
query_gen = QueryGenerator(sql_dict)
for row in ws.iter_rows(name_row[0].row + 1, ws.max_row):
#for row in ws.iter_rows(18, 23):
    #Break on empty row
    if row[0].value is None:
        break
    #Load the data of the row
    for i in range(len(row)):
        for exl in exl_list:
            if i == exl.column:
                exl.data = row[i].value

    #Construct the sql_dict out of the exl_list
    for exl in exl_list:
        for opperation in exl.opperation_list:
            try:
                opperation(exl)
            except Exception as e:

                print("Opperation {} on exl {} failed: ".format(
                    opperation.__name__, exl.name))
                print(e)
                query_gen.skip_row = True
        exl.write()

    if not query_gen.skip_row:
        query_gen.generate(sql_dict)
        query_gen.reset()
    else:
        print('skipping row: ', row[0].row)
        query_gen.failed_queries += 1
        query_gen.skip_row = False
    print('')


print("{}/{} queries successful".format(
query_gen.valid_queries, query_gen.valid_queries + query_gen.failed_queries))

if args["verbose"]:
    commit = input('Commit changes?')
    if commit == 'y':
        conn.commit()
conn.close()
